"""Автоматические генераторы config.yaml для corpus-builder.

Несколько стратегий:
  1. seed_crawl — обходим стартовые URL, собираем ссылки на глубину 1
  2. from_csv — пакетная загрузка URL из CSV/TSV/Excel
  3. from_bookmarks — импорт из bookmark-файла браузера (HTML/Netscape format)
  4. from_github_topics — поиск репозиториев по topic/language
  5. from_stackexchange_tags — топ вопросов по тегам через SE API
  6. discover_sitemaps — чтение sitemap.xml сайтов для массового сбора URL

Каждый генератор возвращает список объектов-источников, которые затем
объединяются и сохраняются как config.yaml.

Использование как CLI:
    python -m corpus_builder.config_generator seed-crawl \\
        --seeds https://habr.com/ru/hub/electronics/ \\
        --max-urls 100 \\
        --output config.generated.yaml

    python -m corpus_builder.config_generator from-csv sources.csv -o config.yaml
    python -m corpus_builder.config_generator from-bookmarks bookmarks.html -o config.yaml
    python -m corpus_builder.config_generator from-github --topic kicad --language Python -o config.yaml
    python -m corpus_builder.config_generator from-stackexchange --site electronics --tag kicad -o config.yaml
    python -m corpus_builder.config_generator from-sitemaps https://habr.com/sitemap.xml -o config.yaml
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin, urlparse

import requests
import yaml
from bs4 import BeautifulSoup

from .models import SourceItem


# ============================================================
# Детектор типа источника по URL
# ============================================================

def detect_source_type(url: str) -> str:
    """Эвристически определить тип источника по URL."""
    parsed = urlparse(url)
    host = parsed.netloc.lower()
    path = parsed.path.lower()

    if host.endswith("github.com") and path.count("/") >= 2:
        return "github_repo"

    if host.endswith("stackexchange.com") or host in (
        "stackoverflow.com", "serverfault.com", "superuser.com",
        "electronics.stackexchange.com", "mathoverflow.net",
    ):
        if "/questions/" in path:
            return "stackexchange"

    if path.endswith(".pdf"):
        return "pdf"

    # По умолчанию — HTML
    return "html"


def make_source(url: str, source_type: str | None = None,
                categories: list[str] | None = None,
                include_files: list[str] | None = None) -> dict:
    """Создать dict-источник для config.yaml."""
    st = source_type or detect_source_type(url)
    item: dict[str, Any] = {"url": url, "type": st}
    if categories:
        item["categories"] = categories
    if include_files and st == "github_repo":
        item["include_files"] = include_files
    return item


# ============================================================
# Стратегия 1: seed_crawl — обход стартовых URL
# ============================================================

def seed_crawl(seeds: list[str], max_urls_per_seed: int = 50,
               same_domain_only: bool = True,
               user_agent: str = "CorpusBuilder/0.2 (config-gen)") -> list[dict]:
    """Обойти стартовые URL и собрать ссылки на одной странице.

    Не делает глубокий обход — только то, что лежит на самих seed-страницах.
    Это безопасно и вежливо: 1 запрос на seed.
    """
    sources: list[dict] = []
    for seed in seeds:
        try:
            r = requests.get(seed, timeout=20,
                             headers={"User-Agent": user_agent})
            r.raise_for_status()
        except Exception as e:
            print(f"  WARN: seed {seed} failed: {e}", file=sys.stderr)
            continue

        soup = BeautifulSoup(r.text, "html.parser")
        seed_domain = urlparse(seed).netloc
        seen_in_seed: set[str] = set()

        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("#") or href.startswith("mailto:") or href.startswith("javascript:"):
                continue
            full_url = urljoin(seed, href)
            parsed = urlparse(full_url)
            # Нормализуем: убираем фрагмент, оставляем scheme+host+path+query
            clean = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
            if parsed.query:
                clean += "?" + parsed.query

            if same_domain_only and parsed.netloc != seed_domain:
                continue
            if clean in seen_in_seed:
                continue
            seen_in_seed.add(clean)

            # Сам seed тоже добавляем
            sources.append(make_source(clean))
            if len(seen_in_seed) >= max_urls_per_seed:
                break

        # Сам seed — тоже источник
        sources.append(make_source(seed))
        print(f"  seed {seed}: collected {len(seen_in_seed)} URLs")

    # Дедуплицируем
    unique: dict[str, dict] = {}
    for s in sources:
        if s["url"] not in unique:
            unique[s["url"]] = s
    return list(unique.values())


# ============================================================
# Стратегия 1b: seed_crawl_depth — BFS-обход с указанной глубиной
# ============================================================

def seed_crawl_depth(
    seed: str,
    depth: int = 0,
    max_urls: int = 100,
    same_domain_only: bool = True,
    include_subdomains: bool = False,
    user_agent: str = "CorpusBuilder/0.2 (config-gen)",
    request_timeout: int = 20,
    request_delay: float = 1.0,
    on_progress: "Callable[[int, int, str], None] | None" = None,
    should_stop: "Callable[[], bool] | None" = None,
) -> list[dict]:
    """BFS-обход от seed на указанную глубину.

    depth=0 — только сам seed (1 источник).
    depth=1 — seed + все ссылки со страницы seed.
    depth=2 — seed + ссылки + ссылки со ссылок. И т.д.

    Возвращает list[dict] готовых источников для config.yaml.

    Параметры:
        seed: стартовый URL
        depth: глубина обхода (0..N)
        max_urls: лимит общего числа собранных URL (защита от взрыва)
        same_domain_only: только same-domain ссылки
        include_subdomains: если True, разрешает blog.example.com для example.com
        on_progress(current, total_estimate, message): callback для GUI
        should_stop(): если возвращает True — останавливаемся
    """
    import time

    if depth < 0:
        depth = 0
    if depth > 1000:
        depth = 1000  # защита от абсурда

    seed_domain = urlparse(seed).netloc
    # Нормализуем seed
    seed_parsed = urlparse(seed)
    seed_clean = f"{seed_parsed.scheme}://{seed_parsed.netloc}{seed_parsed.path}"
    if seed_parsed.query:
        seed_clean += "?" + seed_parsed.query

    visited: set[str] = set()
    sources: list[dict] = []
    # Очередь: (url, current_depth)
    queue: list[tuple[str, int]] = [(seed_clean, 0)]
    sources.append(make_source(seed_clean))
    visited.add(seed_clean)

    def is_allowed_domain(url: str) -> bool:
        if not same_domain_only:
            return True
        parsed = urlparse(url)
        if include_subdomains:
            # example.com совпадает с blog.example.com
            return parsed.netloc == seed_domain or parsed.netloc.endswith("." + seed_domain)
        return parsed.netloc == seed_domain

    while queue:
        if should_stop and should_stop():
            break
        if len(sources) >= max_urls:
            break

        current_url, current_depth = queue.pop(0)
        if current_depth >= depth:
            # Не идём глубже — но сам URL уже в sources
            continue

        if on_progress:
            on_progress(len(sources), max_urls, f"depth={current_depth+1}: {current_url[:80]}")

        try:
            r = requests.get(
                current_url, timeout=request_timeout,
                headers={"User-Agent": user_agent},
            )
            r.raise_for_status()
        except Exception as e:
            if on_progress:
                on_progress(len(sources), max_urls, f"skip {current_url[:60]}: {str(e)[:40]}")
            continue

        soup = BeautifulSoup(r.text, "html.parser")
        new_count = 0
        for a in soup.find_all("a", href=True):
            if len(sources) >= max_urls:
                break
            href = a["href"]
            if href.startswith(("#", "mailto:", "javascript:", "tel:", "data:")):
                continue
            full_url = urljoin(current_url, href)
            parsed = urlparse(full_url)
            if parsed.scheme not in ("http", "https"):
                continue
            clean = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
            if parsed.query:
                clean += "?" + parsed.query
            if clean in visited:
                continue
            if not is_allowed_domain(clean):
                continue
            visited.add(clean)
            sources.append(make_source(clean))
            new_count += 1
            # Добавляем в очередь, если ещё есть глубина
            if current_depth + 1 < depth:
                queue.append((clean, current_depth + 1))

        # Вежливость: задержка между запросами на один домен
        if request_delay > 0:
            time.sleep(request_delay)

    if on_progress:
        on_progress(len(sources), max_urls, "done")
    return sources


# ============================================================
# Стратегия 2b: from_excel — чтение xlsx/xls/csv с колонками url, depth, categories?
# ============================================================

def from_excel(
    path: str | Path,
    url_column: str = "url",
    depth_column: str = "depth",
    categories_column: str | None = "categories",
    sheet_name: str | int | None = 0,
) -> list[tuple[str, int, list[str]]]:
    """Читать URL + глубину из Excel/CSV файла.

    Возвращает list кортежей (url, depth, categories).
    Формат: первая строка — заголовки.
    Колонка categories — опциональная, список через запятую.

    Поддерживаемые форматы (по расширению файла):
        .xlsx — openpyxl
        .xls  — xlrd
        .csv  — встроенный csv (с авто-определением разделителя)
    """
    path = Path(path)
    ext = path.suffix.lower()

    rows: list[dict] = []

    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name or 0]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            wb.close()
            raise ValueError(f"Empty Excel file: {path}")
        headers = [str(h).strip().lower() if h else "" for h in headers]
        for row in rows_iter:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            rows.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                        for i in range(len(headers))})
        wb.close()
    elif ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(sheet_name if isinstance(sheet_name, int) else 0)
        headers = [str(ws.cell_value(0, c)).strip().lower() for c in range(ws.ncols)]
        for r in range(1, ws.nrows):
            row = {headers[c]: str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)}
            if not any(v for v in row.values()):
                continue
            rows.append(row)
    elif ext == ".csv":
        import csv as _csv
        with open(path, "r", encoding="utf-8", newline="") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = _csv.Sniffer().sniff(sample, delimiters=",\t;|")
                reader = _csv.DictReader(f, dialect=dialect)
            except _csv.Error:
                reader = _csv.DictReader(f, delimiter=",")
            for row in reader:
                # В CSV с дубликатами колонок значения могут быть списками
                cleaned = {}
                for k, v in row.items():
                    key = (k or "").strip().lower()
                    if isinstance(v, list):
                        v = v[0] if v else ""
                    cleaned[key] = (v or "").strip()
                rows.append(cleaned)
    else:
        raise ValueError(f"Unsupported file extension: {ext}. Use .xlsx, .xls, or .csv")

    # Приводим к нужному формату
    result: list[tuple[str, int, list[str]]] = []
    for row in rows:
        url = (row.get(url_column.lower()) or "").strip()
        if not url:
            continue
        if not url.startswith(("http://", "https://")):
            # Пропускаем не-URL
            continue
        depth_raw = (row.get(depth_column.lower()) or "0").strip()
        try:
            depth = int(float(depth_raw)) if depth_raw else 0
        except (ValueError, TypeError):
            depth = 0
        depth = max(0, min(1000, depth))  # ограничение 0..1000
        cats: list[str] = []
        if categories_column:
            cats_str = (row.get(categories_column.lower()) or "").strip()
            if cats_str:
                # Поддерживаем разные разделители: запятая, точка с запятой, пайп
                import re
                parts = re.split(r"[,;|]", cats_str)
                cats = [p.strip() for p in parts if p.strip()]
        result.append((url, depth, cats))
    print(f"Loaded {len(result)} sources from {path}")
    return result


def crawl_excel_with_depth(
    path: str | Path,
    max_total_urls: int = 5000,
    on_progress: "Callable[[int, int, str], None] | None" = None,
    should_stop: "Callable[[], bool] | None" = None,
) -> list[dict]:
    """Полный пайплайн: прочитать Excel → для каждой строки сделать seed_crawl_depth.

    Возвращает list[dict] готовых источников для config.yaml.
    """
    rows = from_excel(path)
    if not rows:
        return []

    all_sources: list[dict] = []
    seen_urls: set[str] = set()
    total = len(rows)

    for i, (url, depth, cats) in enumerate(rows):
        if should_stop and should_stop():
            break
        if on_progress:
            on_progress(i, total, f"[{i+1}/{total}] depth={depth}: {url[:60]}")

        # Сам URL добавляем всегда (даже если depth=0)
        if url not in seen_urls:
            source = make_source(url, categories=cats or None)
            all_sources.append(source)
            seen_urls.add(url)

        # Если depth > 0 — делаем BFS-обход
        if depth > 0:
            # Лимит: не более max(50, depth * 50) URL на один seed
            per_seed_max = min(max(50, depth * 50), 1000)
            new_sources = seed_crawl_depth(
                seed=url,
                depth=depth,
                max_urls=per_seed_max,
                on_progress=on_progress,
                should_stop=should_stop,
            )
            for s in new_sources:
                if s["url"] not in seen_urls:
                    if cats:
                        s["categories"] = list(s.get("categories") or []) + cats
                    all_sources.append(s)
                    seen_urls.add(s["url"])

        if on_progress:
            on_progress(i + 1, total, f"[{i+1}/{total}] done, total={len(all_sources)}")

    if on_progress:
        on_progress(total, total, f"done: {len(all_sources)} URLs total")
    return all_sources


# ============================================================
# Стратегия 2: from_csv — пакетная загрузка
# ============================================================

def from_csv(path: str | Path, url_column: str = "url",
             type_column: str | None = "type",
             categories_column: str | None = "categories",
             delimiter: str = ",") -> list[dict]:
    """Читать URL из CSV/TSV.

    Ожидаемые колонки (минимум — url):
        url,type,categories,include_files
        https://...,html,"electronics,pcb",
        https://...,pdf,datasheet,
    """
    sources: list[dict] = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        # Если delimiter не задан, пробуем авто-определение
        if delimiter == "auto":
            sample = f.read(2048)
            f.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            reader = csv.DictReader(f, dialect=dialect)
        else:
            reader = csv.DictReader(f, delimiter=delimiter)

        if url_column not in (reader.fieldnames or []):
            raise ValueError(
                f"Column '{url_column}' not found in CSV. "
                f"Available: {reader.fieldnames}"
            )

        for row in reader:
            url = (row.get(url_column) or "").strip()
            if not url:
                continue
            st = (row.get(type_column) or "").strip() if type_column else None
            cats_str = (row.get(categories_column) or "").strip() if categories_column else ""
            cats = [c.strip() for c in cats_str.split(",") if c.strip()] if cats_str else []
            inc = row.get("include_files")
            include_files = (
                [s.strip() for s in inc.split("|") if s.strip()]
                if inc and inc.strip() else None
            )
            sources.append(make_source(url, st, cats, include_files))
    print(f"Loaded {len(sources)} sources from {path}")
    return sources


# ============================================================
# Стратегия 3: from_bookmarks — импорт из браузера
# ============================================================

def from_bookmarks(path: str | Path) -> list[dict]:
    """Импорт закладок из HTML-файла Netscape Bookmark format.

    Экспорт из Chrome: chrome://bookmarks → ⋮ → Экспорт закладок
    Экспорт из Firefox: Ctrl+Shift+O → Импорт и резервные копии → Экспорт в HTML
    """
    sources: list[dict] = []
    with open(path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f, "html.parser")

    folder_categories: list[str] = []

    def walk(node, depth: int = 0):
        for child in node.find_all(["dt", "h3"], recursive=False):
            if child.name == "h3":
                # Папка — добавляем в categories
                folder_categories.append(child.get_text().strip())
                # Идём вглубь
                next_dl = child.find_next_sibling("dl")
                if next_dl:
                    walk(next_dl, depth + 1)
                folder_categories.pop()
            elif child.name == "dt":
                a = child.find("a")
                if a and a.get("href"):
                    url = a["href"]
                    # Пропускаем javascript: и mailto:
                    if url.startswith(("javascript:", "mailto:", "about:")):
                        continue
                    cats = list(folder_categories) if folder_categories else []
                    sources.append(make_source(url, categories=cats or None))

    # В Netscape format всё обёрнуто в <dl>
    root_dl = soup.find("dl")
    if root_dl:
        walk(root_dl)

    # Дедуплицируем, сохраняя категории из самой глубокой папки
    seen: dict[str, dict] = {}
    for s in sources:
        if s["url"] not in seen:
            seen[s["url"]] = s
    print(f"Imported {len(seen)} bookmarks from {path}")
    return list(seen.values())


# ============================================================
# Стратегия 4: from_github_topics — поиск репозиториев
# ============================================================

def from_github_topics(topics: list[str], language: str | None = None,
                       sort: str = "stars", per_page: int = 30,
                       max_repos: int = 100) -> list[dict]:
    """Найти GitHub-репозитории по topic и/или языку через GitHub Search API.

    Лимиты без токена: 10 запросов/мин. С токеном — 30/мин.
    """
    token = os.environ.get("GITHUB_TOKEN", "")
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"

    sources: list[dict] = []
    for topic in topics:
        page = 1
        while len([s for s in sources if topic in (s.get("categories") or [])]) < max_repos:
            params: dict[str, Any] = {
                "q": f"topic:{topic}" + (f" language:{language}" if language else ""),
                "sort": sort,
                "order": "desc",
                "per_page": per_page,
                "page": page,
            }
            try:
                r = requests.get(
                    "https://api.github.com/search/repositories",
                    params=params, headers=headers, timeout=15,
                )
                if r.status_code == 403:
                    print("  WARN: GitHub rate limit hit. Wait 60s or use GITHUB_TOKEN env.",
                          file=sys.stderr)
                    break
                r.raise_for_status()
                data = r.json()
            except Exception as e:
                print(f"  WARN: GitHub search failed for topic={topic} page={page}: {e}",
                      file=sys.stderr)
                break

            items = data.get("items") or []
            if not items:
                break

            for item in items:
                repo_url = item.get("html_url")
                if not repo_url:
                    continue
                sources.append(make_source(
                    repo_url,
                    source_type="github_repo",
                    categories=[topic, f"stars:{item.get('stargazers_count', 0)}"],
                    include_files=["*.md", "*.kicad_sch", "*.kicad_pcb", "*.csv", "*.dcm", "*.lib"],
                ))
                if len([s for s in sources if topic in (s.get("categories") or [])]) >= max_repos:
                    break

            if len(items) < per_page:
                break
            page += 1
            if page > 10:  # GitHub ограничивает 1000 результатов
                break

    # Дедуплицируем
    seen: dict[str, dict] = {}
    for s in sources:
        if s["url"] not in seen:
            seen[s["url"]] = s
    print(f"Found {len(seen)} repos for topics={topics}, language={language}")
    return list(seen.values())


# ============================================================
# Стратегия 5: from_stackexchange_tags — топ вопросов
# ============================================================

def from_stackexchange_tags(site: str = "electronics", tags: list[str] | None = None,
                            max_questions: int = 100,
                            min_score: int = 5) -> list[dict]:
    """Найти топ-вопросы на StackExchange по тегам.

    API: /questions?site=...&tagged=tag1;tag2&sort=votes&min=5
    """
    api_key = os.environ.get("STACKEXCHANGE_KEY", "")
    base = "https://api.stackexchange.com/2.3"
    sources: list[dict] = []

    # Если теги не заданы — берём популярные по умолчанию
    if not tags:
        tags = ["kicad", "pcb", "schematic", "embedded", "microcontroller"]

    for tag in tags:
        page = 1
        collected = 0
        while collected < max_questions:
            params = {
                "site": site,
                "tagged": tag,
                "sort": "votes",
                "order": "desc",
                "min": min_score,
                "pagesize": 100,
                "page": page,
                "filter": "default",
            }
            if api_key:
                params["key"] = api_key

            try:
                r = requests.get(f"{base}/questions", params=params, timeout=15)
                if r.status_code == 429:
                    print("  WARN: SE rate limit. Wait 30s.", file=sys.stderr)
                    break
                r.raise_for_status()
                data = r.json()
            except Exception as e:
                print(f"  WARN: SE questions failed for tag={tag} page={page}: {e}",
                      file=sys.stderr)
                break

            items = data.get("items") or []
            if not items:
                break

            for q in items:
                qid = q.get("question_id")
                if not qid:
                    continue
                # URL вида https://electronics.stackexchange.com/questions/{id}
                link = q.get("link") or f"https://{site}.stackexchange.com/questions/{qid}"
                sources.append(make_source(
                    link,
                    source_type="stackexchange",
                    categories=[tag, f"score:{q.get('score', 0)}"],
                ))
                collected += 1
                if collected >= max_questions:
                    break

            if len(items) < 100:
                break
            page += 1
            if page > 10:
                break

    seen: dict[str, dict] = {}
    for s in sources:
        if s["url"] not in seen:
            seen[s["url"]] = s
    print(f"Found {len(seen)} questions on {site} for tags={tags}")
    return list(seen.values())


# ============================================================
# Стратегия 6: from_sitemaps — чтение sitemap.xml
# ============================================================

def from_sitemaps(sitemap_urls: list[str], max_urls_per_site: int = 1000,
                  url_filter: str | None = None) -> list[dict]:
    """Извлечь URL из sitemap.xml сайтов.

    Поддерживает обычные и индексные sitemap'ы (recursively).
    Опциональный url_filter — regex-паттерн для фильтрации URL.
    """
    pattern = re.compile(url_filter) if url_filter else None
    sources: list[dict] = []
    visited: set[str] = set()
    queue: list[str] = list(sitemap_urls)

    while queue:
        sitemap_url = queue.pop(0)
        if sitemap_url in visited:
            continue
        visited.add(sitemap_url)

        try:
            r = requests.get(sitemap_url, timeout=20,
                             headers={"User-Agent": "CorpusBuilder/0.2 (sitemap)"})
            r.raise_for_status()
        except Exception as e:
            print(f"  WARN: sitemap {sitemap_url} failed: {e}", file=sys.stderr)
            continue

        # Парсим как XML
        soup = BeautifulSoup(r.text, "xml")
        # Индексный sitemap содержит <sitemap><loc>...</loc></sitemap>
        sub_sitemaps = soup.find_all("sitemap")
        if sub_sitemaps:
            for s in sub_sitemaps:
                loc = s.find("loc")
                if loc and loc.text:
                    queue.append(loc.text.strip())
            continue

        # Обычный sitemap содержит <url><loc>...</loc></url>
        urls = soup.find_all("url")
        collected = 0
        for u in urls:
            loc = u.find("loc")
            if not loc or not loc.text:
                continue
            url = loc.text.strip()
            if pattern and not pattern.search(url):
                continue
            sources.append(make_source(url))
            collected += 1
            if collected >= max_urls_per_site:
                break

        print(f"  sitemap {sitemap_url}: collected {collected} URLs")

    seen: dict[str, dict] = {}
    for s in sources:
        if s["url"] not in seen:
            seen[s["url"]] = s
    print(f"Found {len(seen)} URLs across {len(visited)} sitemaps")
    return list(seen.values())


# ============================================================
# Сборка config.yaml
# ============================================================

DEFAULT_TEMPLATE = {
    "output": {
        "corpus_file": "corpus_output/raw_corpus.jsonl",
        "download_dir": "downloaded_files",
        "error_log": "corpus_output/errors.jsonl",
        "state_file": "corpus_output/state.json",
        "log_file": "corpus_output/crawl.log",
        "max_file_size_mb": 50,
        "request_delay": 2,
        "request_timeout": 30,
        "user_agent": "CorpusBuilder/0.2 (research)",
    },
    "crawlers": {
        "html": {
            "extract_mode": "trafilatura",
            "download_images": True,
            "image_extensions": ["svg", "png", "jpg", "jpeg", "webp"],
            "download_files_ext": ["pdf", "kicad_sch", "kicad_pcb", "zip", "sch", "brd"],
        },
        "pdf": {
            "ocr_enabled": True,
            "ocr_lang": "rus+eng",
            "ocr_min_chars_per_page": 50,
            "image_min_width": 300,
            "image_min_height": 200,
            "extract_tables": False,
        },
        "github": {
            "token_env": "GITHUB_TOKEN",
            "branch": None,
            "include_files": ["*.md", "*.kicad_sch", "*.kicad_pcb", "*.csv", "*.dcm", "*.lib"],
        },
        "stackexchange": {
            "api_key_env": "STACKEXCHANGE_KEY",
            "site": "electronics",
        },
    },
    "quality": {
        "min_chars": 200,
        "max_chars": 200_000,
        "max_non_alpha_ratio": 0.30,
        "max_dup_line_ratio": 0.50,
        "language": "bilingual",
        "languages_allowed": ["ru", "en"],
    },
    "dedup": {
        "exact": True,
        "minhash": True,
        "minhash_num_perm": 128,
        "minhash_threshold": 0.85,
        "dedup_images": True,
    },
    "pipeline": {
        "resume": True,
        "save_checkpoint_every": 50,
        "progress_bar": True,
    },
}


def build_config(sources: list[dict], output_path: str | Path,
                 template: dict | None = None) -> None:
    """Собрать config.yaml из источников и записать в файл."""
    config = dict(template or DEFAULT_TEMPLATE)
    config["sources"] = sources

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        # Заголовок-комментарий
        f.write("# Auto-generated config.yaml for corpus-builder\n")
        f.write(f"# Generated by corpus_builder.config_generator\n")
        f.write(f"# Sources: {len(sources)}\n\n")
        yaml.dump(config, f, default_flow_style=False, sort_keys=False,
                  allow_unicode=True, indent=2)
    print(f"Written {len(sources)} sources to {output_path}")


def merge_sources(config_files: list[str]) -> list[dict]:
    """Умное объединение sources из нескольких config.yaml с дедупликацией.

    Дедупликация по:
      1. Прямой URL — точное совпадение
      2. Канонизированный URL — удаление utm_*, сортировка query, приведение scheme
         (https://example.com/page?id=1&utm_source=email == https://example.com/page?id=1)
      3. URL без trailing slash — example.com/page == example.com/page/

    Если URL повторяется:
      - Берётся запись из первого файла (в порядке аргументов)
      - Категории из дубликатов сливаются в первую запись
    """
    from .text_utils import canonical_url

    sources: list[dict] = []
    seen_exact: dict[str, dict] = {}       # URL -> source
    seen_canonical: dict[str, dict] = {}    # canonical URL -> source
    duplicates_found = 0

    for f in config_files:
        with open(f, "r", encoding="utf-8") as fh:
            cfg = yaml.safe_load(fh) or {}
        for s in cfg.get("sources") or []:
            url = s.get("url", "")
            if not url:
                continue

            # Проверка 1: точное совпадение URL
            if url in seen_exact:
                duplicates_found += 1
                # Сливаем категории из дубликата
                existing = seen_exact[url]
                existing_cats = set(existing.get("categories") or [])
                new_cats = set(s.get("categories") or [])
                merged_cats = list(existing_cats | new_cats)
                if merged_cats:
                    existing["categories"] = merged_cats
                continue

            # Проверка 2: канонизированный URL
            canon = canonical_url(url)
            # Нормализация: убираем trailing slash для сравнения
            canon_key = canon.rstrip("/").lower()
            if canon_key in seen_canonical:
                duplicates_found += 1
                existing = seen_canonical[canon_key]
                # Сливаем категории
                existing_cats = set(existing.get("categories") or [])
                new_cats = set(s.get("categories") or [])
                merged_cats = list(existing_cats | new_cats)
                if merged_cats:
                    existing["categories"] = merged_cats
                continue

            # Уникальный URL — добавляем
            seen_exact[url] = s
            seen_canonical[canon_key] = s
            sources.append(s)

    print(f"Merged {len(sources)} unique sources from {len(config_files)} files "
          f"({duplicates_found} duplicates removed)")
    return sources


def merge_sources_with_stats(config_files: list[str]) -> tuple[list[dict], dict]:
    """Умное объединение с подробной статистикой.

    Возвращает (sources, stats):
      sources: list[dict] уникальных источников
      stats: {
        total_input: int,
        total_output: int,
        duplicates_removed: int,
        by_file: {filename: count},
      }
    """
    from .text_utils import canonical_url

    sources: list[dict] = []
    seen_exact: dict[str, dict] = {}
    seen_canonical: dict[str, dict] = {}
    total_input = 0
    duplicates_found = 0
    by_file: dict[str, int] = {}

    for f in config_files:
        file_count = 0
        with open(f, "r", encoding="utf-8") as fh:
            cfg = yaml.safe_load(fh) or {}
        for s in cfg.get("sources") or []:
            total_input += 1
            file_count += 1
            url = s.get("url", "")
            if not url:
                continue

            is_dup = False
            # Проверка 1: точное совпадение
            if url in seen_exact:
                is_dup = True
                existing = seen_exact[url]
            else:
                # Проверка 2: канонизированный URL
                canon = canonical_url(url)
                canon_key = canon.rstrip("/").lower()
                if canon_key in seen_canonical:
                    is_dup = True
                    existing = seen_canonical[canon_key]

            if is_dup:
                duplicates_found += 1
                # Сливаем категории
                existing_cats = set(existing.get("categories") or [])
                new_cats = set(s.get("categories") or [])
                merged_cats = list(existing_cats | new_cats)
                if merged_cats:
                    existing["categories"] = merged_cats
                continue

            seen_exact[url] = s
            canon = canonical_url(url)
            canon_key = canon.rstrip("/").lower()
            seen_canonical[canon_key] = s
            sources.append(s)

        by_file[Path(f).name] = file_count

    stats = {
        "total_input": total_input,
        "total_output": len(sources),
        "duplicates_removed": duplicates_found,
        "by_file": by_file,
    }
    print(f"Merge stats: {stats}")
    return sources, stats


def merge_sources_into_config(config_files: list[str], output_path: str | Path) -> None:
    """Объединить несколько config.yaml в один и записать результат."""
    sources = merge_sources(config_files)
    build_config(sources, output_path)


# ============================================================
# Стратегия 7: from_wikipedia — поиск статей по категориям
# ============================================================

def from_wikipedia(
    categories: list[str],
    lang: str = "en",
    max_articles: int = 50,
    depth: int = 0,
) -> list[dict]:
    """Найти статьи Wikipedia по категориям через MediaWiki API.

    Использует API /api.php?action=query&list=categorymembers
    для получения списка статей в заданных категориях.

    Параметры:
        categories: список категорий Wikipedia (например, ["Electronics", "Printed circuit boards"])
        lang: языковой код Wikipedia (en, ru, de, fr, ...)
        max_articles: максимум статей на категорию
        depth: глубина обхода подкатегорий (0 = только прямые статьи)

    Возвращает list[dict] готовых источников.
    """
    import requests as _req

    sources: list[dict] = []
    seen_urls: set[str] = set()

    api_url = f"https://{lang}.wikipedia.org/w/api.php"

    for category in categories:
        # Убираем "Category:" префикс если есть
        cat_name = category.replace("Category:", "").replace("Категория:", "").strip()

        # Получаем статьи в категории
        params = {
            "action": "query",
            "list": "categorymembers",
            "cmtitle": f"Category:{cat_name}",
            "cmlimit": min(max_articles, 500),
            "cmtype": "page",  # только статьи, не подкатегории
            "cmprop": "title",
            "format": "json",
        }

        try:
            r = _req.get(api_url, params=params, timeout=15,
                         headers={"User-Agent": "CorpusBuilder/0.2 (wikipedia-search)"})
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"  WARN: Wikipedia search failed for '{cat_name}': {e}", file=sys.stderr)
            continue

        members = (data.get("query", {}) or {}).get("categorymembers") or []
        if not members:
            print(f"  Wikipedia category '{cat_name}' ({lang}): no articles found", file=sys.stderr)
            continue

        count = 0
        for member in members:
            title = member.get("title", "")
            if not title or title.startswith(("Category:", "File:", "Template:", "Wikipedia:")):
                continue

            # Формируем URL статьи
            from urllib.parse import quote
            article_url = f"https://{lang}.wikipedia.org/wiki/{quote(title.replace(' ', '_'))}"

            if article_url in seen_urls:
                continue
            seen_urls.add(article_url)

            sources.append(make_source(
                article_url,
                categories=[f"wikipedia:{lang}", f"category:{cat_name.lower()}"],
            ))
            count += 1
            if count >= max_articles:
                break

        print(f"  Wikipedia '{cat_name}' ({lang}): found {count} articles")

    # Также ищем в подкатегориях если depth > 0
    if depth > 0:
        for category in categories:
            cat_name = category.replace("Category:", "").replace("Категория:", "").strip()
            subcats = _get_wikipedia_subcategories(cat_name, lang, api_url)
            for subcat in subcats[:10]:  # ограничиваем глубину
                sub_sources = from_wikipedia(
                    [subcat], lang=lang, max_articles=max_articles, depth=depth - 1
                )
                for s in sub_sources:
                    if s["url"] not in seen_urls:
                        seen_urls.add(s["url"])
                        sources.append(s)

    print(f"Wikipedia: found {len(sources)} articles across {len(categories)} categories")
    return sources


def _get_wikipedia_subcategories(category: str, lang: str, api_url: str) -> list[str]:
    """Получить список подкатегорий для категории Wikipedia."""
    import requests as _req
    params = {
        "action": "query",
        "list": "categorymembers",
        "cmtitle": f"Category:{category}",
        "cmlimit": 50,
        "cmtype": "subcat",
        "cmprop": "title",
        "format": "json",
    }
    try:
        r = _req.get(api_url, params=params, timeout=15,
                     headers={"User-Agent": "CorpusBuilder/0.2 (wikipedia-search)"})
        r.raise_for_status()
        data = r.json()
        members = (data.get("query", {}) or {}).get("categorymembers") or []
        # Возвращаем полные названия категорий
        return [m.get("title", "").replace("Category:", "") for m in members if m.get("title")]
    except Exception:
        return []


# ============================================================
# Мультиязычный поиск Wikipedia — несколько языков сразу
# ============================================================

def from_wikipedia_multi(
    categories: list[str],
    languages: list[str] = None,
    max_articles: int = 50,
    depth: int = 0,
) -> list[dict]:
    """Поиск статей Wikipedia на нескольких языках одновременно.

    Параметры:
        categories: список категорий (одни и те же для всех языков)
        languages: список языковых кодов (например, ["en", "ru"])
        max_articles: максимум статей на категорию на каждый язык
        depth: глубина обхода подкатегорий

    Возвращает list[dict] с дедупликацией URL (разные языки = разные URL).
    """
    if languages is None:
        languages = ["en"]

    all_sources: list[dict] = []
    seen_urls: set[str] = set()
    stats: dict[str, int] = {}

    for lang in languages:
        print(f"Wikipedia: searching lang={lang}, categories={categories}")
        sources = from_wikipedia(
            categories=categories,
            lang=lang,
            max_articles=max_articles,
            depth=depth,
        )
        count = 0
        for s in sources:
            if s["url"] not in seen_urls:
                seen_urls.add(s["url"])
                all_sources.append(s)
                count += 1
        stats[lang] = count
        print(f"  Wikipedia ({lang}): {count} unique articles")

    print(f"Wikipedia multi-lang: {len(all_sources)} total articles across {len(languages)} languages")
    print(f"  By language: {stats}")
    return all_sources


# ============================================================
# Шаблон Excel для пользователя
# ============================================================

def save_template_xlsx(output_path: str | Path) -> str:
    """Создать .xlsx-шаблон с правильными заголовками и примером.

    Возвращает путь к созданному файлу.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sources"

    # Заголовки
    headers = ["url", "depth", "categories"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="007ACC")
        cell.alignment = Alignment(horizontal="center")

    # Примеры
    examples = [
        ("https://habr.com/ru/hub/electronics/", 1, "electronics,pcb"),
        ("https://github.com/user/awesome-kicad", 0, "kicad,library"),
        ("https://electronics.stackexchange.com/questions/322180", 0, "electronics,qa"),
        ("https://example.com/datasheet.pdf", 0, "datasheet"),
    ]
    for i, (url, depth, cats) in enumerate(examples, start=2):
        ws.cell(row=i, column=1, value=url)
        ws.cell(row=i, column=2, value=depth)
        ws.cell(row=i, column=3, value=cats)

    # Подсказки
    ws.cell(row=7, column=1, value="# url — обязательная колонка (с http:// или https://)")
    ws.cell(row=8, column=1, value="# depth — 0..1000, глубина BFS-обхода от этого URL")
    ws.cell(row=9, column=1, value="# categories — опционально, список через запятую")
    for r in range(7, 10):
        ws.cell(row=r, column=1).font = Font(italic=True, color="808080")

    # Ширина колонок
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 30

    # Закрепить заголовок
    ws.freeze_panes = "A2"

    wb.save(str(output_path))
    print(f"Template saved: {output_path}")
    return str(output_path)


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        prog="config_generator",
        description="Auto-generate config.yaml for corpus-builder",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    # seed-crawl
    p_seed = subparsers.add_parser("seed-crawl",
                                   help="Crawl seed URLs to discover more")
    p_seed.add_argument("--seeds", nargs="+", required=True,
                        help="Seed URLs to start from")
    p_seed.add_argument("--max-urls", type=int, default=50,
                        help="Max URLs per seed (default: 50)")
    p_seed.add_argument("--cross-domain", action="store_true",
                        help="Allow URLs from any domain (default: same-domain only)")
    p_seed.add_argument("-o", "--output", default="config.generated.yaml",
                        help="Output config.yaml path")

    # from-csv
    p_csv = subparsers.add_parser("from-csv",
                                  help="Load URLs from CSV/TSV file")
    p_csv.add_argument("file", help="CSV/TSV file path")
    p_csv.add_argument("--url-column", default="url")
    p_csv.add_argument("--type-column", default="type")
    p_csv.add_argument("--categories-column", default="categories")
    p_csv.add_argument("--delimiter", default="auto",
                      help="Delimiter: comma, tab, semicolon, pipe, or auto")
    p_csv.add_argument("-o", "--output", default="config.generated.yaml")

    # from-bookmarks
    p_bm = subparsers.add_parser("from-bookmarks",
                                 help="Import URLs from browser bookmarks HTML")
    p_bm.add_argument("file", help="Bookmarks HTML file (Netscape format)")
    p_bm.add_argument("-o", "--output", default="config.generated.yaml")

    # from-github
    p_gh = subparsers.add_parser("from-github",
                                 help="Search GitHub repos by topic/language")
    p_gh.add_argument("--topics", nargs="+", required=True,
                      help="GitHub topics (e.g., kicad pcb embedded)")
    p_gh.add_argument("--language", help="Filter by programming language")
    p_gh.add_argument("--max-repos", type=int, default=100,
                      help="Max repos per topic (default: 100)")
    p_gh.add_argument("-o", "--output", default="config.generated.yaml")

    # from-stackexchange
    p_se = subparsers.add_parser("from-stackexchange",
                                 help="Find top questions on StackExchange by tags")
    p_se.add_argument("--site", default="electronics",
                      help="SE site name (default: electronics)")
    p_se.add_argument("--tags", nargs="+",
                      help="Tags to search (default: kicad pcb schematic embedded microcontroller)")
    p_se.add_argument("--max-questions", type=int, default=100,
                      help="Max questions per tag (default: 100)")
    p_se.add_argument("--min-score", type=int, default=5,
                      help="Min question score (default: 5)")
    p_se.add_argument("-o", "--output", default="config.generated.yaml")

    # from-sitemaps
    p_sm = subparsers.add_parser("from-sitemaps",
                                 help="Extract URLs from sitemap.xml files")
    p_sm.add_argument("--sitemaps", nargs="+", required=True,
                      help="Sitemap URLs (supports index sitemaps recursively)")
    p_sm.add_argument("--max-urls-per-site", type=int, default=1000)
    p_sm.add_argument("--filter", help="Regex to filter URLs")
    p_sm.add_argument("-o", "--output", default="config.generated.yaml")

    # from-wikipedia — поиск статей Wikipedia по категориям
    p_wiki = subparsers.add_parser("from-wikipedia",
                                    help="Find Wikipedia articles by categories")
    p_wiki.add_argument("--categories", nargs="+", required=True,
                         help="Wikipedia categories (e.g., Electronics 'Printed circuit boards')")
    p_wiki.add_argument("--lang", default="en",
                         help="Wikipedia language code (en, ru, de, fr, ...)")
    p_wiki.add_argument("--max-articles", type=int, default=50,
                        help="Max articles per category (default: 50)")
    p_wiki.add_argument("--depth", type=int, default=0,
                        help="Subcategory recursion depth (default: 0)")
    p_wiki.add_argument("-o", "--output", default="config.generated.yaml")

    # merge — объединить несколько конфигов
    p_merge = subparsers.add_parser("merge",
                                    help="Merge multiple config.yaml files")
    p_merge.add_argument("files", nargs="+", help="config.yaml files to merge")
    p_merge.add_argument("-o", "--output", default="config.merged.yaml")

    args = parser.parse_args()

    # Делимитер для CSV
    delim_map = {"comma": ",", "tab": "\t", "semicolon": ";", "pipe": "|", "auto": "auto"}
    if hasattr(args, "delimiter"):
        args.delimiter = delim_map.get(args.delimiter, args.delimiter)

    # Запуск соответствующего генератора
    if args.command == "seed-crawl":
        sources = seed_crawl(
            args.seeds,
            max_urls_per_seed=args.max_urls,
            same_domain_only=not args.cross_domain,
        )
    elif args.command == "from-csv":
        sources = from_csv(
            args.file,
            url_column=args.url_column,
            type_column=args.type_column,
            categories_column=args.categories_column,
            delimiter=args.delimiter,
        )
    elif args.command == "from-bookmarks":
        sources = from_bookmarks(args.file)
    elif args.command == "from-github":
        sources = from_github_topics(
            args.topics,
            language=args.language,
            max_repos=args.max_repos,
        )
    elif args.command == "from-stackexchange":
        sources = from_stackexchange_tags(
            site=args.site,
            tags=args.tags,
            max_questions=args.max_questions,
            min_score=args.min_score,
        )
    elif args.command == "from-sitemaps":
        sources = from_sitemaps(
            args.sitemaps,
            max_urls_per_site=args.max_urls_per_site,
            url_filter=args.filter,
        )
    elif args.command == "from-wikipedia":
        sources = from_wikipedia(
            categories=args.categories,
            lang=args.lang,
            max_articles=args.max_articles,
            depth=args.depth,
        )
    elif args.command == "merge":
        # Объединить sources из нескольких config.yaml с дедупликацией
        sources = merge_sources(args.files)
        # Дедупликация уже сделана в merge_sources
        print(f"Merged {len(sources)} unique sources from {len(args.files)} files")
    else:
        parser.print_help()
        sys.exit(1)

    if not sources:
        print("ERROR: no sources collected", file=sys.stderr)
        sys.exit(1)

    build_config(sources, args.output)


if __name__ == "__main__":
    main()
