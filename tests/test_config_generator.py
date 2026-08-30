"""Тесты на config_generator."""

import yaml

import corpus_builder.config_generator as cg
from corpus_builder.config_generator import (
    build_config,
    detect_source_type,
    from_csv,
    from_excel,
    make_source,
    save_template_xlsx,
    seed_crawl_depth,
)


def test_detect_source_type_github():
    assert detect_source_type("https://github.com/user/repo") == "github_repo"


def test_detect_source_type_stackexchange():
    assert detect_source_type(
        "https://electronics.stackexchange.com/questions/12345/title"
    ) == "stackexchange"


def test_detect_source_type_pdf():
    assert detect_source_type("https://example.com/datasheet.pdf") == "pdf"


def test_detect_source_type_html_default():
    assert detect_source_type("https://habr.com/ru/articles/123/") == "html"


def test_make_source_minimal():
    s = make_source("https://example.com/page")
    assert s["url"] == "https://example.com/page"
    assert s["type"] == "html"


def test_make_source_with_categories():
    s = make_source("https://example.com/page", categories=["electronics", "pcb"])
    assert s["categories"] == ["electronics", "pcb"]


def test_make_source_github_with_include_files():
    s = make_source(
        "https://github.com/user/repo",
        source_type="github_repo",
        include_files=["*.md", "*.kicad_sch"],
    )
    assert s["type"] == "github_repo"
    assert s["include_files"] == ["*.md", "*.kicad_sch"]


def test_from_csv_basic(tmp_path):
    csv_file = tmp_path / "sources.csv"
    csv_file.write_text(
        "url,type,categories\n"
        "https://habr.com/ru/articles/1/,html,electronics\n"
        "https://example.com/doc.pdf,pdf,datasheet\n"
        "https://github.com/user/repo,github_repo,kicad\n",
        encoding="utf-8",
    )
    sources = from_csv(csv_file, delimiter=",")
    assert len(sources) == 3
    assert sources[0]["url"] == "https://habr.com/ru/articles/1/"
    assert sources[0]["type"] == "html"
    assert sources[0]["categories"] == ["electronics"]


def test_from_csv_auto_delimiter_tsv(tmp_path):
    csv_file = tmp_path / "sources.tsv"
    csv_file.write_text(
        "url\ttype\n"
        "https://example.com/a\thtml\n"
        "https://example.com/b\thtml\n",
        encoding="utf-8",
    )
    sources = from_csv(csv_file, delimiter="auto")
    assert len(sources) == 2


def test_from_csv_missing_url_column(tmp_path):
    csv_file = tmp_path / "bad.csv"
    csv_file.write_text("name,description\nfoo,bar\n", encoding="utf-8")
    try:
        from_csv(csv_file, delimiter=",")
        assert False, "Should have raised"
    except ValueError as e:
        assert "url" in str(e)


def test_build_config_writes_yaml(tmp_path):
    sources = [
        {"url": "https://example.com/1", "type": "html"},
        {"url": "https://example.com/2.pdf", "type": "pdf"},
    ]
    out = tmp_path / "config.yaml"
    build_config(sources, out)
    assert out.exists()
    with open(out, encoding="utf-8") as f:
        # Пропускаем первые строки-комментарии
        content = "\n".join(l for l in f.read().splitlines() if not l.startswith("#"))
    cfg = yaml.safe_load(content)
    assert "sources" in cfg
    assert len(cfg["sources"]) == 2
    assert "output" in cfg
    assert "crawlers" in cfg
    assert "quality" in cfg
    assert "dedup" in cfg


def test_build_config_preserves_categories(tmp_path):
    sources = [
        {"url": "https://example.com/1", "type": "html",
         "categories": ["electronics", "pcb"]},
    ]
    out = tmp_path / "config.yaml"
    build_config(sources, out)
    with open(out, encoding="utf-8") as f:
        content = "\n".join(l for l in f.read().splitlines() if not l.startswith("#"))
    cfg = yaml.safe_load(content)
    assert cfg["sources"][0]["categories"] == ["electronics", "pcb"]


def test_merge_sources_into_config_dedup(tmp_path):
    """Если один и тот же URL в двух конфигах, должен остаться один."""
    f1 = tmp_path / "c1.yaml"
    f2 = tmp_path / "c2.yaml"
    f1.write_text(
        "sources:\n"
        "  - url: 'https://example.com/1'\n"
        "    type: html\n"
        "    categories: [a]\n",
        encoding="utf-8",
    )
    f2.write_text(
        "sources:\n"
        "  - url: 'https://example.com/1'\n"
        "    type: html\n"
        "    categories: [b]\n"
        "  - url: 'https://example.com/2'\n"
        "    type: pdf\n",
        encoding="utf-8",
    )
    sources = cg.merge_sources([str(f1), str(f2)])
    assert len(sources) == 2
    urls = [s["url"] for s in sources]
    assert "https://example.com/1" in urls
    assert "https://example.com/2" in urls


# ============================================================
# Тесты на seed_crawl_depth
# ============================================================

def test_seed_crawl_depth_zero():
    """depth=0 возвращает только сам URL (без сетевых запросов)."""
    sources = seed_crawl_depth("https://example.com/", depth=0)
    assert len(sources) == 1
    assert sources[0]["url"] == "https://example.com/"


def test_seed_crawl_depth_clamping():
    """Глубина > 1000 ограничивается до 1000."""
    # Это проверка без сетевого запроса, depth слишком большой но функция не падает
    sources = seed_crawl_depth("https://example.com/", depth=2000, max_urls=1)
    assert len(sources) == 1  # только seed


def test_seed_crawl_depth_negative():
    """Отрицательная глубина → 0."""
    sources = seed_crawl_depth("https://example.com/", depth=-5)
    assert len(sources) == 1


def test_seed_crawl_depth_progress_callback():
    """on_progress вызывается."""
    progress_calls = []
    def on_progress(current, total, msg):
        progress_calls.append((current, total, msg))
    seed_crawl_depth("https://example.com/", depth=0, on_progress=on_progress)
    # Хотя бы один вызов (финальный)
    assert len(progress_calls) >= 1


# ============================================================
# Тесты на from_excel
# ============================================================

def test_from_excel_csv(tmp_path):
    csv_file = tmp_path / "sources.csv"
    csv_file.write_text(
        "url,depth,categories\n"
        "https://habr.com/ru/hub/electronics/,1,electronics;pcb\n"
        "https://github.com/user/repo,0,kicad\n"
        "https://example.com/datasheet.pdf,0,\n",
        encoding="utf-8",
    )
    rows = from_excel(csv_file)
    assert len(rows) == 3
    assert rows[0][0] == "https://habr.com/ru/hub/electronics/"
    assert rows[0][1] == 1
    # categories через ";" внутри одного CSV-поля
    cats0 = rows[0][2]
    assert "electronics" in cats0
    assert "pcb" in cats0
    assert rows[1][1] == 0
    assert rows[2][2] == []  # пустые категории


def test_from_excel_xlsx(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["url", "depth", "categories"])
    ws.append(["https://example.com/page1", 2, "test,docs"])
    ws.append(["https://example.com/page2", 0, ""])
    ws.append(["https://github.com/user/repo", 0, "kicad"])
    xlsx_file = tmp_path / "sources.xlsx"
    wb.save(str(xlsx_file))

    rows = from_excel(xlsx_file)
    assert len(rows) == 3
    assert rows[0][0] == "https://example.com/page1"
    assert rows[0][1] == 2
    assert "test" in rows[0][2]
    assert rows[1][1] == 0
    assert rows[2][2] == ["kicad"]


def test_from_excel_invalid_url_skipped(tmp_path):
    csv_file = tmp_path / "bad.csv"
    csv_file.write_text(
        "url,depth\n"
        "not-a-url,1\n"
        "https://valid.example.com,1\n"
        "ftp://unsupported.scheme/,0\n"
        ",0\n",
        encoding="utf-8",
    )
    rows = from_excel(csv_file)
    assert len(rows) == 1  # только валидный https-URL
    assert rows[0][0] == "https://valid.example.com"


def test_from_excel_depth_clamping(tmp_path):
    csv_file = tmp_path / "depth.csv"
    csv_file.write_text(
        "url,depth\n"
        "https://example.com/a,5000\n"
        "https://example.com/b,-3\n"
        "https://example.com/c,not_a_number\n",
        encoding="utf-8",
    )
    rows = from_excel(csv_file)
    assert len(rows) == 3
    # 5000 clamped to 1000
    assert rows[0][1] == 1000
    # -3 clamped to 0
    assert rows[1][1] == 0
    # not_a_number → 0
    assert rows[2][1] == 0


def test_from_excel_unsupported_extension(tmp_path):
    bad_file = tmp_path / "sources.txt"
    bad_file.write_text("url\ndepth\n", encoding="utf-8")
    try:
        from_excel(bad_file)
        assert False, "Should have raised"
    except ValueError as e:
        assert "txt" in str(e) or "Unsupported" in str(e)


# ============================================================
# Тесты на save_template_xlsx
# ============================================================

def test_save_template_xlsx(tmp_path):
    out = tmp_path / "template.xlsx"
    result = save_template_xlsx(out)
    assert out.exists()
    assert out.stat().st_size > 0

    # Проверить содержимое
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
    assert headers == ["url", "depth", "categories"]
    # В строке 2 должен быть первый пример
    assert "https://" in (ws.cell(row=2, column=1).value or "")
    wb.close()


# ============================================================
# Тесты на build_config с реальным набором источников
# ============================================================

def test_build_config_after_excel(tmp_path):
    csv_file = tmp_path / "sources.csv"
    csv_file.write_text(
        "url,depth,categories\n"
        "https://habr.com/ru/hub/electronics/,1,electronics\n"
        "https://github.com/user/repo,0,kicad\n",
        encoding="utf-8",
    )
    rows = from_excel(csv_file)
    sources = [make_source(url, categories=cats or None) for url, _, cats in rows]
    out = tmp_path / "config.yaml"
    build_config(sources, out)
    with open(out, encoding="utf-8") as f:
        content = "\n".join(l for l in f.read().splitlines() if not l.startswith("#"))
    cfg = yaml.safe_load(content)
    assert len(cfg["sources"]) == 2
    assert cfg["sources"][0]["categories"] == ["electronics"]
